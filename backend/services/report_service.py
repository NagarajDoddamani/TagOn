from io import BytesIO, StringIO
import csv
from datetime import datetime, date, timezone
from sqlalchemy.orm import Session
from models.order import Order
from models.payment import Payment
from models.user import User
from models.product import Product


class ReportService:
    def __init__(self, db: Session):
        self.db = db

    def generate(self, report_type: str, fmt: str, **filters):
        data_methods = {
            "orders": self._orders_data,
            "payments": self._payments_data,
            "revenue": self._revenue_data,
            "customers": self._customers_data,
            "products": self._products_data,
        }
        titles = {
            "orders": "Orders Report",
            "payments": "Payments Report",
            "revenue": "Revenue Report",
            "customers": "Customers Report",
            "products": "Products Report",
        }
        headers, rows = data_methods[report_type](**filters)
        title = titles[report_type]
        filename = f"{report_type}-report-{date.today().isoformat()}"

        if fmt == "csv":
            return self._to_csv(headers, rows, filename), "text/csv"
        elif fmt == "xlsx":
            return self._to_xlsx(title, headers, rows, filename), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif fmt == "pdf":
            return self._to_pdf(title, headers, rows, filename), "application/pdf"

    # ------------------------------------------------------------------ #
    #  Data gathering
    # ------------------------------------------------------------------ #

    def _orders_data(self, start_date=None, end_date=None, status=None):
        q = self.db.query(Order).options(
            *[]  # joinedload removed – we do explicit joins below
        )
        if start_date:
            q = q.filter(Order.created_at >= start_date)
        if end_date:
            q = q.filter(Order.created_at <= end_date)
        if status:
            q = q.filter(Order.order_status == status)
        orders = q.order_by(Order.created_at.desc()).all()

        headers = ["Order ID", "Customer", "Product", "Quantity", "Total",
                    "Order Status", "Payment Status", "Created At"]
        rows = []
        for o in orders:
            cust = self.db.query(User).filter(User.id == o.customer_id).first()
            rows.append([
                str(o.id), cust.name if cust else "N/A",
                o.product.name if o.product else "N/A",
                o.quantity, o.total_amount,
                o.order_status, o.payment_status,
                o.created_at.isoformat() if o.created_at else "",
            ])
        return headers, rows

    def _payments_data(self, start_date=None, end_date=None):
        q = self.db.query(Payment).join(Order, Payment.order_id == Order.id)
        if start_date:
            q = q.filter(Payment.created_at >= start_date)
        if end_date:
            q = q.filter(Payment.created_at <= end_date)
        payments = q.order_by(Payment.created_at.desc()).all()

        headers = ["Payment ID", "Order ID", "Transaction ID",
                    "Verification Status", "Created At"]
        rows = []
        for p in payments:
            rows.append([
                str(p.id), str(p.order_id),
                p.transaction_id or "",
                p.verification_status,
                p.created_at.isoformat() if p.created_at else "",
            ])
        return headers, rows

    def _revenue_data(self, start_date=None, end_date=None):
        q = self.db.query(Order).filter(
            Order.order_status.notin_(["cancelled", "archived"])
        )
        if start_date:
            q = q.filter(Order.created_at >= start_date)
        if end_date:
            q = q.filter(Order.created_at <= end_date)
        orders = q.order_by(Order.created_at.asc()).all()

        headers = ["Date", "Orders Count", "Revenue"]
        daily = {}
        for o in orders:
            d = o.created_at.date().isoformat() if o.created_at else "unknown"
            entry = daily.setdefault(d, {"count": 0, "revenue": 0})
            entry["count"] += 1
            entry["revenue"] += o.total_amount

        rows = []
        running_total = 0
        for d in sorted(daily):
            running_total += daily[d]["revenue"]
            rows.append([d, daily[d]["count"], round(daily[d]["revenue"], 2)])
        rows.append(["TOTAL", "", round(running_total, 2)])
        return headers, rows

    def _customers_data(self, start_date=None, end_date=None):
        q = self.db.query(User).filter(User.role == "customer")
        if start_date:
            q = q.filter(User.created_at >= start_date)
        if end_date:
            q = q.filter(User.created_at <= end_date)
        customers = q.order_by(User.created_at.desc()).all()

        headers = ["Customer ID", "Name", "Email", "Phone", "Status", "Registered At"]
        rows = []
        for c in customers:
            rows.append([
                str(c.id), c.name, c.email, c.phone,
                c.status, c.created_at.isoformat() if c.created_at else "",
            ])
        return headers, rows

    def _products_data(self, category_id=None, status=None):
        q = self.db.query(Product).filter(Product.is_deleted == False)
        if category_id:
            q = q.filter(Product.category_id == category_id)
        if status:
            q = q.filter(Product.status == status)
        products = q.order_by(Product.created_at.desc()).all()

        headers = ["Product ID", "Name", "Type", "Base Price",
                    "Status", "Category", "Created At"]
        rows = []
        for p in products:
            rows.append([
                str(p.id), p.name, p.product_type, p.base_price,
                p.status, p.category.name if p.category else "N/A",
                p.created_at.isoformat() if p.created_at else "",
            ])
        return headers, rows

    # ------------------------------------------------------------------ #
    #  Format writers
    # ------------------------------------------------------------------ #

    def _to_csv(self, headers, rows, filename):
        buf = StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        w.writerows(rows)
        content = buf.getvalue().encode("utf-8-sig")
        return content, filename + ".csv"

    def _to_xlsx(self, title, headers, rows, filename):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read(), filename + ".xlsx"

    def _to_pdf(self, title, headers, rows, filename):
        from fpdf import FPDF

        pdf = FPDF(orientation="L" if len(headers) > 6 else "P")
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 12, title, new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(4)

        col_width = (pdf.w - 20) / len(headers)
        col_width = min(col_width, 50)

        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255, 255, 255)
        for h in headers:
            pdf.cell(col_width, 10, h[:30], border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(0, 0, 0)
        fill = False
        for row in rows:
            max_h = 8
            x_start = pdf.get_x()
            y_start = pdf.get_y()

            if y_start + max_h > pdf.h - 20:
                pdf.add_page()
                pdf.set_font("Helvetica", "B", 9)
                pdf.set_fill_color(31, 78, 121)
                pdf.set_text_color(255, 255, 255)
                for h in headers:
                    pdf.cell(col_width, 10, h[:30], border=1, align="C", fill=True)
                pdf.ln()
                pdf.set_font("Helvetica", "", 8)
                pdf.set_text_color(0, 0, 0)

            if fill:
                pdf.set_fill_color(240, 240, 240)
            else:
                pdf.set_fill_color(255, 255, 255)

            for i, val in enumerate(row):
                text = str(val) if val is not None else ""
                pdf.cell(col_width, 8, text[:30], border=1, align="C", fill=True)
            pdf.ln()
            fill = not fill

        buf = BytesIO()
        pdf.output(buf)
        buf.seek(0)
        return buf.read(), filename + ".pdf"
